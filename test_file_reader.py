import fitz
from docx import Document
from openpyxl import load_workbook


def read_pdf(path):
    text = ""
    doc = fitz.open(path)

    for page in doc:
        text += page.get_text()

    return text


def read_docx(path):
    doc = Document(path)
    return "\n".join([p.text for p in doc.paragraphs])


def read_excel(path):
    wb = load_workbook(path)
    sheet = wb.active

    rows = []

    for row in sheet.iter_rows(values_only=True):
        rows.append(" ".join([str(cell) for cell in row if cell]))

    return "\n".join(rows)


def read_txt(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


if __name__ == "__main__":
    files = [
        "sample.pdf",
        "sample.docx",
        "sample.xlsx",
        "sample.txt"
    ]

    for file in files:

        print(f"\n{'=' * 50}")
        print(f"Reading: {file}")

        if file.endswith(".pdf"):
            print(read_pdf(file)[:500])

        elif file.endswith(".docx"):
            print(read_docx(file)[:500])

        elif file.endswith(".xlsx"):
            print(read_excel(file)[:500])

        elif file.endswith(".txt"):
            print(read_txt(file)[:500])